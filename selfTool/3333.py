from openpyxl import load_workbook


def is_last_column(file_path, sheet_name, cell):
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    cell_column = str(cell.column_letter)
    last_column = sheet.max_column
    return cell_column == last_column


# Usage
print(is_last_column(r'C:\Users\wangze\Desktop\sheet1.xlsx', 'result 1', 'A1'))